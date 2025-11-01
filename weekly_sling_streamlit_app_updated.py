# app.py
# Streamlit app: Generate Weekly Sling Payroll Workbook with guardrails
# UPDATED: UHL included, no unpaid break, 24h guardrail → 6h

import io
import re
from datetime import datetime
from typing import Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st

# ---------------- Defaults (updated to match new requirements) ----------------
THRESH_HOURS_DEFAULT = 13.0
GUARD_24H_PAYS = 6.0  # NEW: if worked > 24h, pay fixed 6.0h
DEFAULT_EXCLUDE_REGEX = r"solid\s*fire|\badministration\b"  # UHL NOW INCLUDED
DEFAULT_UNPAID_BREAK_HOURS = 0.0  # NO BREAK DEDUCTION by default
DEFAULT_UNPAID_BREAK_THRESHOLD = 6.0


# ---------------- Helper functions ----------------
def to_date(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce")


def latest_week_window(dates_series: pd.Series) -> Tuple[pd.Timestamp, pd.Timestamp]:
    dates = to_date(dates_series).dropna()
    if dates.empty:
        raise RuntimeError("Export has no valid dates.")
    latest = dates.max().normalize()
    # Saturday anchor (Mon=0..Sun=6 -> Sat=5)
    shift = (latest.weekday() - 5) % 7
    week_start = (latest - pd.Timedelta(days=shift)).normalize()
    week_end = week_start + pd.Timedelta(days=6)
    return week_start, week_end


def parse_excel_time(val) -> Optional[int]:
    """Return minutes since midnight (0..1439) or None."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    if isinstance(val, pd.Timestamp):
        return val.hour * 60 + val.minute
    if isinstance(val, (int, float)) and 0 <= val <= 1:
        mins = int(round(val * 24 * 60))
        return max(0, min(mins, 24 * 60 - 1))
    s = str(val).strip()
    if s == "" or s.lower() in {"nan", "none", "-"}:
        return None
    s = s.replace(".,", ":").replace(".", ":").upper()
    for fmt in ("%I:%M %p", "%I %p", "%H:%M", "%H.%M", "%H:%M:%S"):
        try:
            t = datetime.strptime(s, fmt)
            return t.hour * 60 + t.minute
        except Exception:
            pass
    if s.isdigit() and (3 <= len(s) <= 4):
        hh = int(s[:-2])
        mm = int(s[-2:])
        if 0 <= hh < 24 and 0 <= mm < 60:
            return hh * 60 + mm
    return None


def parse_hours(s) -> float:
    if pd.isna(s):
        return np.nan
    st = str(s).replace(",", "").strip()
    if st in {"", "-", "nan", "none", "None"}:
        return np.nan
    try:
        return float(st)
    except Exception:
        try:
            h, m = st.split(":")
            return float(h) + float(m) / 60.0
        except Exception:
            return np.nan


def full_day_for_dow(dow: str) -> float:
    if dow in ("Monday", "Tuesday", "Wednesday", "Thursday"):
        return 8.0
    if dow == "Friday":
        return 7.0
    return 0.0


def ok_time(x) -> bool:
    return (x is not None) and not (isinstance(x, float) and np.isnan(x))


# Leave / training / holiday detection
VTH_REGEXES = [
    re.compile(r"\bannual\s*leave\b", re.I),
    re.compile(r"\bvacation\b", re.I),
    re.compile(r"\btraining\b", re.I),
    re.compile(r"\b(induction|course)\b", re.I),
    re.compile(
        r"\b(bank\s*holiday|public\s*holiday|company\s*holiday|stat(?:utory)?\s*holiday|holiday)\b", re.I),
    re.compile(r"(?<!\w)A/?L(?!\w)", re.I),
    re.compile(r"(?<!\w)B/?H(?!\w)", re.I),
    re.compile(r"(?<!\w)P/?H(?!\w)", re.I),
]


def detect_vth(row: pd.Series) -> bool:
    text = " ".join(str(row.get(c, ""))
                    for c in ("POSITIONS", "LOCATIONS", "STATUS", "NOTES"))
    return any(rx.search(text or "") for rx in VTH_REGEXES)


def default_sched(row: pd.Series) -> Tuple[Optional[int], Optional[int]]:
    """
    Position-based defaults:
      Special employees: 07:00–15:30 Mon–Thu, 07:00–14:00 Fri
      Anthony Grogan:    07:30–16:00 Mon–Thu, 07:30–15:00 Fri
      Firestoppers:      07:30–16:00 Mon–Thu, 07:30–15:00 Fri
      Painters:          07:00–15:30 Mon–Thu, 07:00–14:00 Fri
    """
    dow = row["Day"]
    employee = (row.get("Employee") or "").strip()
    pos = (row.get("Position") or "").lower()

    # Special employee schedules: 7:00-15:30 Mon-Thu, 7:00-14:00 Fri
    special_employees_7_1530 = [
        "Bruno Henrique Ragasini Souza",
        "Gary Hughes",
        "Chris Hughes",
        "Vadim Mihalas",
    ]

    # Anthony Grogan: 7:30-16:00 Mon-Thu, 7:30-15:00 Fri
    special_employees_730_16 = [
        "Anthony Grogan",
    ]

    if employee in special_employees_7_1530:
        start = 7 * 60  # 7:00 AM
        if dow in ("Monday", "Tuesday", "Wednesday", "Thursday"):
            end = 15 * 60 + 30  # 3:30 PM
        elif dow == "Friday":
            end = 14 * 60  # 2:00 PM
        else:
            return (None, None)
        return (start, end)

    if employee in special_employees_730_16:
        start = 7 * 60 + 30  # 7:30 AM
        if dow in ("Monday", "Tuesday", "Wednesday", "Thursday"):
            end = 16 * 60  # 4:00 PM
        elif dow == "Friday":
            end = 15 * 60  # 3:00 PM
        else:
            return (None, None)
        return (start, end)

    if "firestop" in pos or "fire stopper" in pos:
        start = 7 * 60 + 30
        if dow in ("Monday", "Tuesday", "Wednesday", "Thursday"):
            end = 16 * 60
        elif dow == "Friday":
            end = 15 * 60
        else:
            return (None, None)
        return (start, end)
    if "painter" in pos:
        start = 7 * 60
        if dow in ("Monday", "Tuesday", "Wednesday", "Thursday"):
            end = 15 * 60 + 30
        elif dow == "Friday":
            end = 14 * 60
        else:
            return (None, None)
        return (start, end)
    return (None, None)


def mm_to_hhmm(m: Optional[int]) -> str:
    if not ok_time(m):
        return ""
    m = int(m)
    return f"{m // 60:02d}:{m % 60:02d}"


# ---------------- Core generator ----------------
def generate_weekly_workbook(
    df_raw: pd.DataFrame,
    exclude_regex: str = DEFAULT_EXCLUDE_REGEX,
    thresh_hours: float = THRESH_HOURS_DEFAULT,
    unpaid_break_hours: float = DEFAULT_UNPAID_BREAK_HOURS,
    unpaid_break_threshold: float = DEFAULT_UNPAID_BREAK_THRESHOLD,
) -> Tuple[bytes, pd.DataFrame, Tuple[pd.Timestamp, pd.Timestamp]]:
    """
    Returns (excel_bytes, weekly_totals_df, (week_start, week_end))
    """

    # Normalize headers to canonical set used in pipeline
    expected = {
        "EMPLOYEE": ["EMPLOYEE", "Employee", "Name"],
        "DATE": ["DATE", "Date"],
        "CLOCK IN": ["CLOCK IN", "CLOCK IN\nTIME", "Clock In", "IN", "Time In"],
        "CLOCK OUT": ["CLOCK OUT", "CLOCK OUT\nTIME", "Clock Out", "OUT", "Time Out"],
        "AUTO": ["AUTO\nCLOCK-OUT", "Auto Clock-out", "AUTO CLOCK OUT", "Auto Clock Out", "Auto"],
        "SHIFT DUR": ["SHIFT\nDURATION", "Worked", "Duration", "Hours"],
        "STATUS": ["STATUS", "Status"],
        "NOTES": ["NOTES", "Notes", "Comments"],
        "POSITIONS": ["POSITIONS", "Position"],
        "LOCATIONS": ["LOCATIONS", "Location", "Site"],
    }
    col_map = {}
    for tgt, alts in expected.items():
        for a in alts:
            if a in df_raw.columns:
                col_map[a] = tgt
                break
    df = df_raw.rename(columns=col_map).copy()
    for k in expected.keys():
        if k not in df.columns:
            df[k] = np.nan

    # Week window filter
    week_start, week_end = latest_week_window(df["DATE"])
    in_week = (to_date(df["DATE"]) >= week_start) & (
        to_date(df["DATE"]) <= week_end)
    df = df[in_week].copy()

    # Derived columns
    df["Employee"] = df["EMPLOYEE"].astype(str).str.strip()
    df["Clock In"] = df["CLOCK IN"].astype(str)
    df["Clock Out"] = df["CLOCK OUT"].astype(str)
    df["Auto"] = df["AUTO"].astype(str).str.strip().str.lower().eq("yes")
    df["Worked_raw"] = df["SHIFT DUR"].apply(parse_hours)

    df["cin_min"] = df["CLOCK IN"].apply(parse_excel_time)
    df["cout_min"] = df["CLOCK OUT"].apply(parse_excel_time)
    df["has_clock_in"] = df["Clock In"].apply(lambda s: str(
        s).strip() not in {"", "-", "nan", "none", "None"})
    df["has_clock_out"] = df["Clock Out"].apply(lambda s: str(s).strip() not in {
                                                "", "-", "nan", "none", "None"})

    text_flags = df[["STATUS", "NOTES"]].astype(
        str).agg(" ".join, axis=1).str.lower()
    df["no_clock_out_text"] = text_flags.str.contains("no clock out", na=False)
    df["valid_punch"] = df["has_clock_in"] & (
        df["has_clock_out"] | df["Auto"]) & (~df["no_clock_out_text"])

    df["Date"] = to_date(df["DATE"]).dt.date
    df["Day"] = to_date(df["DATE"]).dt.day_name()
    df["Position"] = df["POSITIONS"].astype(str)
    df["Location"] = df["LOCATIONS"].astype(str)

    # Exclusions (Position/Location contains pattern)
    exc_re = re.compile(exclude_regex, re.I) if exclude_regex else None
    if exc_re:
        mask_ex = df["Position"].str.contains(
            exc_re, na=False) | df["Location"].str.contains(exc_re, na=False)
        excluded_rows = df[mask_ex].copy()
        df = df[~mask_ex].copy()
    else:
        excluded_rows = df.iloc[0:0].copy()

    # Schedule-trim
    df["SchedStart_min"], df["SchedEnd_min"] = zip(
        *df.apply(default_sched, axis=1))

    def effective_minutes(row):
        if not row["valid_punch"]:
            return 0.0, "Invalid punch"
        ci, co, ss, se = row["cin_min"], row["cout_min"], row["SchedStart_min"], row["SchedEnd_min"]
        if ok_time(ci) and ok_time(co):
            raw = co - ci
            # If invalid duration, fall back to Worked_raw
            if raw <= 0 or raw >= 24 * 60:
                if pd.notna(row["Worked_raw"]):
                    if ok_time(ss) and ok_time(se):
                        sched_minutes = max(0, int(se) - int(ss))
                        return min(int(round(row["Worked_raw"] * 60)), sched_minutes), "From Worked_raw (invalid clock times)"
                    return int(round(row["Worked_raw"] * 60)), "From Worked_raw (invalid clock times)"
                return 0.0, "Invalid raw duration"
            if ok_time(ss) and ok_time(se):
                start = max(ci, int(ss))
                end = min(co, int(se))
                eff = max(0, end - start)
                return eff, "Trimmed to schedule" if eff < raw else "Within schedule"
            return raw, "No schedule"
        if pd.notna(row["Worked_raw"]):
            if ok_time(ss) and ok_time(se):
                sched_minutes = max(0, int(se) - int(ss))
                return min(int(round(row["Worked_raw"] * 60)), sched_minutes), "Capped by schedule duration"
            return int(round(row["Worked_raw"] * 60)), "From Worked_raw"
        return 0.0, "No times or worked"

    df[["eff_minutes", "eff_note"]] = df.apply(
        lambda r: pd.Series(effective_minutes(r)), axis=1)

    # Guardrails: 13h and no-clockout-without-auto
    df["raw_minutes"] = ((df["cout_min"].fillna(0) - df["cin_min"].fillna(0))).where(
        df["cout_min"].notna() & df["cin_min"].notna(), np.nan
    )
    df["over13_raw"] = (df["raw_minutes"] / 60.0) > float(thresh_hours)
    df["over13_eff"] = (df["eff_minutes"] / 60.0) > float(thresh_hours)
    df["no_clockout_noauto"] = (~df["has_clock_out"]) & (~df["Auto"])
    df["guard_zero"] = df["over13_raw"] | df["over13_eff"] | df["no_clockout_noauto"] | df["no_clock_out_text"]

    # Special 24h rule: if Worked_raw = 24, set to 6h instead of 0
    df["is_24h"] = (df["Worked_raw"] == 24.0)

    # Worked (before unpaid break)
    # Priority: 24h → 6.0, guard_zero → 0.0, else → eff_minutes
    df["Worked_final"] = np.where(
        df["is_24h"], 6.0,
        np.where(df["guard_zero"], 0.0, df["eff_minutes"] / 60.0)
    )

    # Leave / Training / Holiday
    df["VacTrainHoliday"] = df.apply(detect_vth, axis=1)

    # UNPAID BREAK: deduct (strict > threshold), before caps, not on leave/training/holiday
    if unpaid_break_hours and unpaid_break_hours > 0:
        df["Break Applied (h)"] = np.where(
            (df["Worked_final"] > float(unpaid_break_threshold)) & (
                ~df["VacTrainHoliday"]),
            float(unpaid_break_hours),
            0.0,
        )
    else:
        df["Break Applied (h)"] = 0.0

    df["Worked_payable"] = np.maximum(
        0.0, df["Worked_final"] - df["Break Applied (h)"])

    # Daily policy caps (after break applied)
    def policy_paid(row):
        dow = row["Day"]
        if row["VacTrainHoliday"]:
            return full_day_for_dow(dow)
        if dow == "Sunday":
            return 0.0
        if not row["valid_punch"]:
            return 0.0
        w = row["Worked_payable"] if pd.notna(row["Worked_payable"]) else 0.0

        # 24h guardrail already handled in Worked_final (set to 6h)
        # Just apply daily caps
        if dow in ("Monday", "Tuesday", "Wednesday", "Thursday"):
            return min(w, 8.0)
        if dow == "Friday":
            return min(w, 7.0)
        if dow == "Saturday":
            return w
        return 0.0

    df["Policy Paid"] = df.apply(policy_paid, axis=1)

    # ---------- Audit ----------
    audit = df[
        [
            "Employee",
            "Date",
            "Day",
            "Position",
            "Location",
            "Clock In",
            "Clock Out",
            "Auto",
            "Worked_raw",
            "Worked_final",
            "Break Applied (h)",
            "Worked_payable",
            "Policy Paid",
            "VacTrainHoliday",
            "valid_punch",
            "eff_note",
            "guard_zero",
            "is_24h",
            "over13_raw",
            "over13_eff",
            "no_clockout_noauto",
            "SchedStart_min",
            "SchedEnd_min",
        ]
    ].copy()
    audit["Sched Start"] = audit["SchedStart_min"].apply(mm_to_hhmm)
    audit["Sched End"] = audit["SchedEnd_min"].apply(mm_to_hhmm)
    audit.drop(columns=["SchedStart_min", "SchedEnd_min"], inplace=True)
    audit = audit.sort_values(
        ["Employee", "Date", "Day"]).reset_index(drop=True)

    # Spaced Audit (blank row between employees)
    spaced = []
    prev_emp = None
    for _, r in audit.iterrows():
        if prev_emp is not None and r["Employee"] != prev_emp:
            spaced.append({c: np.nan for c in audit.columns})
        spaced.append(r.to_dict())
        prev_emp = r["Employee"]
    audit_spaced = pd.DataFrame(spaced, columns=audit.columns)

    # ---------- Weekly totals ----------
    weekly = (
        audit.groupby("Employee", as_index=False)["Policy Paid"]
        .sum()
        .rename(columns={"Policy Paid": "Sling Total"})
    )

    flags = (
        audit.groupby("Employee")
        .agg(
            **{
                "Any Valid Punch": ("valid_punch", lambda s: bool((s == True).any())),
                "Any Vac/Train/Holiday": ("VacTrainHoliday", lambda s: bool((s == True).any())),
            }
        )
        .reset_index()
    )
    weekly = weekly.merge(flags, on="Employee", how="left")
    # No-carry unless leave present
    weekly.loc[~weekly["Any Valid Punch"] & ~
               weekly["Any Vac/Train/Holiday"], "Sling Total"] = np.nan

    # Add manual entry columns (blank for user to fill)
    weekly["Holiday Hours"] = np.nan
    weekly["OT"] = np.nan
    weekly["Travel"] = np.nan
    weekly["Comments"] = ""
    weekly["Dan Comments"] = ""

    # Net Total = Sling Total + Holiday Hours + OT (will show NaN initially until manual entry)
    weekly["Net Total"] = weekly["Sling Total"] + weekly["Holiday Hours"].fillna(0.0) + weekly["OT"].fillna(0.0)

    # QC tabs
    qc_zeroed = audit[audit["guard_zero"]][
        [
            "Employee",
            "Date",
            "Day",
            "Position",
            "Clock In",
            "Clock Out",
            "Worked_final",
            "Worked_payable",
            "Policy Paid",
            "over13_raw",
            "over13_eff",
            "no_clockout_noauto",
            "eff_note",
        ]
    ]
    qc_trimmed = audit[audit["eff_note"].astype(str).str.contains("Trimmed", na=False)][
        [
            "Employee",
            "Date",
            "Day",
            "Position",
            "Clock In",
            "Clock Out",
            "Worked_final",
            "Worked_payable",
            "Policy Paid",
            "eff_note",
        ]
    ]

    # Guardrails sheet (updated text)
    break_text = f"UNPAID BREAK: deduct {unpaid_break_hours}h when Worked_final > {unpaid_break_threshold}h (strictly over {unpaid_break_threshold}h; before caps; not on leave/training/holiday)." if unpaid_break_hours > 0 else "NO unpaid break deduction."

    guardrails_text = [
        "UHL employees are INCLUDED (not excluded).",
        f"EXCLUSIONS: Remove rows where Position/Location contains 'Solid Fire' or 'Administration'. Regex: {exclude_regex or '(none)'}",
        f"ZERO if daily raw or trimmed > {thresh_hours} hours (UNLESS Worked_raw = 24h, then pay 6h).",
        "ZERO if no clock-out and Auto != Yes, or 'NO CLOCK OUT' flagged.",
        "Schedule-trim by Position/Employee:",
        "  - Special 7:00 group (Bruno Henrique Ragasini Souza, Gary Hughes, Chris Hughes, Vadim Mihalas): 07:00–15:30 Mon–Thu / 14:00 Fri",
        "  - Anthony Grogan: 07:30–16:00 Mon–Thu / 15:00 Fri",
        "  - Firestoppers: 07:30–16:00 Mon–Thu / 15:00 Fri",
        "  - Painters: 07:00–15:30 Mon–Thu / 14:00 Fri",
        "Daily policy: Mon–Thu min(Worked,8), Fri min(Worked,7), Sat full, Sun 0.",
        break_text,
        f"24h GUARDRAIL: if Worked_raw = 24.0 hours exactly, set Worked_final to {GUARD_24H_PAYS}h (instead of 0).",
        "Leave/Training/Holiday: pay full weekday day (Mon–Thu 8h, Fri 7h).",
        "No-carry: blank Sling Total if no valid punches and no leave.",
        "",
        "WEEKLY TOTALS TAB:",
        "  - Sling Total: Auto-calculated from daily Policy Paid totals",
        "  - Holiday Hours, OT, Travel: BLANK (manual entry)",
        "  - Net Total = Sling Total + Holiday Hours + OT (Excel formula)",
        "  - Comments, Dan Comments: BLANK (manual entry)",
        "",
        "HIGHLIGHTING in Audit tabs:",
        "  - BLUE: Clock In time is after Sched Start (late arrival)",
        "  - RED: Clock Out time is before Sched End (early departure)",
        "  - YELLOW: Worked_raw = 24.0 hours (24h guardrail triggered → 6h pay)",
    ]

    # Build workbook in memory
    buffer = io.BytesIO()
    wk_label = f"{week_start.date():%Y-%m-%d}_to_{week_end.date():%Y-%m-%d}"
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        weekly[
            [
                "Employee",
                "Sling Total",
                "Holiday Hours",
                "OT",
                "Net Total",
                "Travel",
                "Any Valid Punch",
                "Any Vac/Train/Holiday",
                "Comments",
                "Dan Comments",
            ]
        ].sort_values("Employee").to_excel(writer, sheet_name="Weekly Totals", index=False)

        # Add Excel formulas to Net Total column
        ws_weekly = writer.sheets["Weekly Totals"]
        # Find Net Total column (column E, index 5)
        for row_idx in range(2, ws_weekly.max_row + 1):  # Start from row 2 (after header)
            # Net Total = Sling Total (B) + Holiday Hours (C) + OT (D)
            # Using IFERROR to handle blanks gracefully
            formula = f"=IFERROR(B{row_idx},0)+IFERROR(C{row_idx},0)+IFERROR(D{row_idx},0)"
            ws_weekly.cell(row=row_idx, column=5).value = formula

        audit_spaced.to_excel(writer, sheet_name="Audit", index=False)
        audit.to_excel(writer, sheet_name="Audit (No Spacing)", index=False)
        qc_trimmed.to_excel(writer, sheet_name="QC - Trimmed", index=False)
        qc_zeroed.to_excel(writer, sheet_name="QC - Zeroed", index=False)
        # Excluded view: show raw if present
        keep_cols = [c for c in ["EMPLOYEE", "DATE", "POSITIONS", "LOCATIONS",
                                 "CLOCK IN", "CLOCK OUT", "SHIFT DUR", "STATUS", "NOTES"] if c in df_raw.columns]
        excluded_view = df_raw[keep_cols].copy(
        ) if keep_cols else df_raw.copy()
        # Only keep actual excluded rows if we computed them
        if exc_re is not None and not excluded_rows.empty:
            excluded_view = excluded_rows[keep_cols].copy(
            ) if keep_cols else excluded_rows.copy()
        excluded_view.to_excel(writer, sheet_name="QC - Excluded", index=False)
        pd.DataFrame({"Guardrails": guardrails_text}).to_excel(
            writer, sheet_name="Guardrails", index=False)

        # Apply conditional formatting to Audit tabs
        from openpyxl.styles import PatternFill

        # Blue fill for late clock-ins, Red fill for early clock-outs, Yellow fill for 24h worked
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for sheet_name in ["Audit", "Audit (No Spacing)"]:
            ws = writer.sheets[sheet_name]

            # Find column indices for Clock In, Clock Out, Sched Start, Sched End, and Worked_raw
            header_row = 1
            clock_in_col = None
            clock_out_col = None
            sched_start_col = None
            sched_end_col = None
            worked_raw_col = None

            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value == "Clock In":
                    clock_in_col = col_idx
                elif cell.value == "Clock Out":
                    clock_out_col = col_idx
                elif cell.value == "Sched Start":
                    sched_start_col = col_idx
                elif cell.value == "Sched End":
                    sched_end_col = col_idx
                elif cell.value == "Worked_raw":
                    worked_raw_col = col_idx

            # Apply highlighting to data rows
            for row_idx in range(2, ws.max_row + 1):
                # Highlight Clock In if after Sched Start (blue)
                if clock_in_col and sched_start_col:
                    clock_in_cell = ws.cell(row=row_idx, column=clock_in_col)
                    sched_start_cell = ws.cell(row=row_idx, column=sched_start_col)

                    # Parse time strings to compare
                    if clock_in_cell.value and sched_start_cell.value:
                        try:
                            # Handle various time formats
                            clock_in_str = str(clock_in_cell.value).strip()
                            sched_start_str = str(sched_start_cell.value).strip()

                            if clock_in_str and sched_start_str and sched_start_str != "":
                                # Parse HH:MM format
                                if ":" in clock_in_str and ":" in sched_start_str:
                                    ci_parts = clock_in_str.split()
                                    ci_time = ci_parts[0] if ci_parts else ""
                                    ss_parts = sched_start_str.split(":")

                                    # Convert to minutes for comparison
                                    if ":" in ci_time:
                                        ci_hm = ci_time.split(":")
                                        ci_min = int(ci_hm[0]) * 60 + int(ci_hm[1])
                                        # Handle AM/PM if present
                                        if len(ci_parts) > 1 and "PM" in ci_parts[1].upper() and int(ci_hm[0]) != 12:
                                            ci_min += 12 * 60
                                        elif len(ci_parts) > 1 and "AM" in ci_parts[1].upper() and int(ci_hm[0]) == 12:
                                            ci_min -= 12 * 60

                                        ss_min = int(ss_parts[0]) * 60 + int(ss_parts[1])

                                        if ci_min > ss_min:
                                            clock_in_cell.fill = blue_fill
                        except Exception:
                            pass  # Skip if parsing fails

                # Highlight Clock Out if before Sched End (red)
                if clock_out_col and sched_end_col:
                    clock_out_cell = ws.cell(row=row_idx, column=clock_out_col)
                    sched_end_cell = ws.cell(row=row_idx, column=sched_end_col)

                    # Parse time strings to compare
                    if clock_out_cell.value and sched_end_cell.value:
                        try:
                            # Handle various time formats
                            clock_out_str = str(clock_out_cell.value).strip()
                            sched_end_str = str(sched_end_cell.value).strip()

                            if clock_out_str and sched_end_str and sched_end_str != "":
                                # Parse HH:MM format
                                if ":" in clock_out_str and ":" in sched_end_str:
                                    co_parts = clock_out_str.split()
                                    co_time = co_parts[0] if co_parts else ""
                                    se_parts = sched_end_str.split(":")

                                    # Convert to minutes for comparison
                                    if ":" in co_time:
                                        co_hm = co_time.split(":")
                                        co_min = int(co_hm[0]) * 60 + int(co_hm[1])
                                        # Handle AM/PM if present
                                        if len(co_parts) > 1 and "PM" in co_parts[1].upper() and int(co_hm[0]) != 12:
                                            co_min += 12 * 60
                                        elif len(co_parts) > 1 and "AM" in co_parts[1].upper() and int(co_hm[0]) == 12:
                                            co_min -= 12 * 60

                                        se_min = int(se_parts[0]) * 60 + int(se_parts[1])

                                        if co_min < se_min:
                                            clock_out_cell.fill = red_fill
                        except Exception:
                            pass  # Skip if parsing fails

                # Highlight Worked_raw if == 24 (yellow)
                if worked_raw_col:
                    worked_raw_cell = ws.cell(row=row_idx, column=worked_raw_col)
                    if worked_raw_cell.value is not None:
                        try:
                            if float(worked_raw_cell.value) == 24.0:
                                worked_raw_cell.fill = yellow_fill
                        except (ValueError, TypeError):
                            pass  # Skip if not a number

    buffer.seek(0)
    return buffer.read(), weekly.sort_values("Employee"), (week_start, week_end)


# ---------------- Streamlit UI ----------------
st.set_page_config(page_title="Weekly Sling Payroll Generator", layout="wide")
st.title("Weekly Sling Payroll Generator (Updated)")
st.caption("UHL Included • No Unpaid Break • 24h Guardrail → 6h")

with st.sidebar:
    st.header("Options")
    exclude_regex = st.text_input("Exclusion regex (Position/Location)", value=DEFAULT_EXCLUDE_REGEX,
                                  help="Rows matching this regex in Position OR Location will be removed. Default excludes Solid Fire and Administration (UHL is INCLUDED).")
    thresh_hours = st.number_input(
        "Daily cap (hours): ZERO if raw/trimmed > cap", value=THRESH_HOURS_DEFAULT, step=0.5, min_value=1.0)
    use_unpaid_break = st.checkbox(
        "Apply unpaid break when Worked_final > threshold", value=False)
    unpaid_break_hours = DEFAULT_UNPAID_BREAK_HOURS if use_unpaid_break else 0.0
    if use_unpaid_break:
        unpaid_break_hours = st.number_input(
            "Unpaid break (hours)", value=0.5, step=0.25, min_value=0.0)
        unpaid_break_threshold = st.number_input(
            "Unpaid break threshold (hours, strict >)", value=DEFAULT_UNPAID_BREAK_THRESHOLD, step=0.25, min_value=0.0)
    else:
        unpaid_break_threshold = DEFAULT_UNPAID_BREAK_THRESHOLD

uploaded = st.file_uploader(
    "Upload Sling export (.xls or .xlsx)", type=["xls", "xlsx"])

if uploaded is not None:
    try:
        # Use engine fallback so .xls works (needs xlrd)
        df_raw = pd.read_excel(uploaded, sheet_name=0)
        excel_bytes, weekly_df, (week_start, week_end) = generate_weekly_workbook(
            df_raw=df_raw,
            exclude_regex=exclude_regex,
            thresh_hours=thresh_hours,
            unpaid_break_hours=unpaid_break_hours,
            unpaid_break_threshold=unpaid_break_threshold,
        )

        st.success(
            f"Workbook created for week {week_start.date():%Y-%m-%d} to {week_end.date():%Y-%m-%d}")

        filename_suffix = "_NO-BREAK" if unpaid_break_hours == 0 else f"_BREAK-{unpaid_break_hours}h"
        st.download_button(
            label="⬇️ Download Weekly Workbook (.xlsx)",
            data=excel_bytes,
            file_name=f"WC_{week_start.date():%Y-%m-%d}_to_{week_end.date():%Y-%m-%d}_MASTER_UHL-INCLUDED_24H-6h{filename_suffix}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.subheader("Weekly Totals Preview")
        st.dataframe(weekly_df, use_container_width=True)

        with st.expander("Show guardrails applied"):
            break_display = f"**UNPAID BREAK:** deduct {unpaid_break_hours}h when Worked_final > {unpaid_break_threshold}h (strictly over {unpaid_break_threshold}h; before caps; not on leave/training/holiday)." if unpaid_break_hours > 0 else "**NO unpaid break deduction.**"

            st.markdown(
                f"""
**Guardrails**

- **UHL employees are INCLUDED** (not excluded).
- **EXCLUSIONS:** Remove rows where Position/Location contains 'Solid Fire' or 'Administration'.
- **ZERO** if daily raw or trimmed > {thresh_hours} hours (UNLESS Worked_raw = 24h, then pay 6h).
- **ZERO** if no clock-out and Auto != Yes, or 'NO CLOCK OUT' flagged.
- **Schedule-trim** by Position/Employee:
  - Special 7:00 group (Bruno Henrique Ragasini Souza, Gary Hughes, Chris Hughes, Vadim Mihalas): 07:00–15:30 Mon–Thu / 14:00 Fri
  - Anthony Grogan: 07:30–16:00 Mon–Thu / 15:00 Fri
  - Firestoppers: 07:30–16:00 Mon–Thu / 15:00 Fri
  - Painters: 07:00–15:30 Mon–Thu / 14:00 Fri
- **Daily policy:** Mon–Thu min(Worked,8), Fri min(Worked,7), Sat full, Sun 0.
- {break_display}
- **24h GUARDRAIL:** if Worked_raw = 24.0 hours exactly, set Worked_final to {GUARD_24H_PAYS}h (instead of 0).
- **Leave/Training/Holiday:** pay full weekday day (Mon–Thu 8h, Fri 7h).
- **No-carry:** blank Sling Total if no valid punches and no leave.

**Weekly Totals Tab:**
- **Sling Total:** Auto-calculated from daily Policy Paid totals
- **Holiday Hours, OT, Travel:** BLANK (manual entry)
- **Net Total:** = Sling Total + Holiday Hours + OT
- **Comments, Dan Comments:** BLANK (manual entry)

**Excel Highlighting in Audit Tabs:**
- 🔵 **BLUE:** Clock In time is after Sched Start (late arrival)
- 🔴 **RED:** Clock Out time is before Sched End (early departure)
- 🟡 **YELLOW:** Worked_raw = 24.0 hours (24h guardrail → pays 6h instead of 0)
                """
            )
    except Exception as e:
        st.error(f"Failed to process file: {e}")
else:
    st.info("Upload a Sling export to begin.")
